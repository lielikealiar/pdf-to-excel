import pdf2image 
from google.cloud import vision 
import openpyxl #to create and write excel files with proper formatting
import openpyxl.styles #for bolding the headers
import sys
import os

#PDF TO IMAGE STRUCTURE
def convert_pdf_to_images(pdf_path):
    if not os.path.exists(pdf_path):
        raise FileNotFoundError(f"PDF file not found: {pdf_path}")
    
    images = pdf2image.convert_from_path(pdf_path) #reads pdf and returns a list where each item is an of one page
    return images

# OCR PROCESSING SETUP
def extract_text_from_image(client, image):
    '''
    We're separating OCR processing into its own function because we'll need to do this for every page. 
    It's cleaner to write the OCR code once and call it multiple times.
    '''
    import io #to convert our image to a byte stream format

    image_bytes = io.BytesIO() #creates an empty byte buffer - think of it as a container that will hold our image data in the format Google expects
    image.save(image_bytes, format='PNG') #saving our image into byte buffer as PNG file
    image_bytes = image_bytes.getvalue() #getvalue extracts the actual byte data from our buffer, leaving variable with only the raw png data
    image_content = vision.Image(content=image_bytes) #wrapping byte_data in google vison's image object

    response = client.text_detection(image_content) #actual ocr call 

    return response

#TEXT STRUCTURE ANALYSIS
def analyze_table_structure(text_data):
    '''
    figures out how the ocr text is organized into rows and columns.
    '''
    extracted_blocks = [] #vision returns text in blocks -- rectangular areas containing related text

    for annotation in  text_data.text_annotations:
        '''
        Google Vision returns multiple types of text detection. 
        text_annotations contains individual text elements with their positions. 
        We loop through each one
        '''
        if annotation.description: #some annotations maybe empty and contain only white spaces
            vertices = annotation.bounding_poly.vertices #vertices contains the coordinates of the rectangular box around the text
            x_coords = [vertex.x for vertex in vertices] #extracting all x coordinate of bounding boxes
            y_coords = [vertex.y for vertex in vertices] #extracting all y coordinates of the box

            left = min(x_coords)
            right = max(x_coords)
            top = min(y_coords)
            bottom = max(y_coords)

            '''
            storing all important information about each text piece:
            what it says and exactly where its positioned.
            '''
            text_block = {
                'text': annotation.description,
                'left': left,
                'right': right,
                'top': top,
                'bottom': bottom
            }

            extracted_blocks.append(text_block)

    return extracted_blocks

#ROW DETECTION
def group_text_into_rows(text_blocks):
    if not text_blocks:
        return[]
    
    '''
    sorting text by vertical position
    '''
    sorted_blocks = sorted(text_blocks, key=lambda x: x['top']) #sorting all text blocks from top to bottom of the page
    rows = [] #to store grouped rows
    current_row = [sorted_blocks[0]] ##we start the first row with the top most text block. each row needs at least one text block
    current_row_y = sorted_blocks[0]['top'] #the vertical position of current row -- used to decide if other text blocks belong to the same row

    '''
    Group text by rows
    '''
    for block in sorted_blocks[1:]: #we skip first block as we have already added to the current block
        y_difference = abs(block['top'] - current_row_y) #we see how far vertically this block is from the current row
        
        if y_difference <= 5: #if less than 10 pixels vertically, its the same row
            current_row.append(block)
        else:
            rows.append(current_row)
            current_row = [block] #starting a new row as this block as first item
            current_row_y = block['top'] #updating the row position with refernce to new block vertical position

    rows.append(current_row)
    return rows
    
#COLUMN DETECTION WITHIN ROWS
def sort_row_into_columns(row_blocks):
    '''
    Within each row we arrange text from left to right to create proper columns
    '''
    sorted_row = sorted(row_blocks, key=lambda x: x['left']) #sorting by horizontal position -- in natural column order
    return sorted_row #returning row's text in proper left to right order

#TABLE ORGANIZATION FUNCTION
def organize_into_table(text_blocks):
    '''
    row grouping and column grouping creating final table structure
    '''
    rows = group_text_into_rows(text_blocks)

    organized_table = []

    for row in rows:
        sorted_row = sort_row_into_columns(row)
        row_texts = [block['text'].strip() for block in sorted_row] #extract just the text content from each block and remove white spaces
        organized_table.append(row_texts)

    return organized_table

#EXCEL WRITING FUNCTION
def write_table_to_excel(table_data, output_filename):
    from openpyxl import workbook #workbook is the main class for creating excel files
    wb = openpyxl.Workbook() #craeting a blank excel workbook programmatically
    ws = wb.active #we get the active worksheet where we will write our table; not calling just referncing to the worksheet

    if table_data:
        headers = table_data[0]

        #writing headers
        for col_num, headers in enumerate(headers, 1): #gives both header text and column number starting from 1
            cell = ws.cell(row=1, column=col_num) #getting specific cellin row 1 and current col number to write the header file here
            cell.value = headers #setting cell's content to header text
            cell.font = openpyxl.styles.Font(bold=True) #setting styles for header text

        #writing data rows
        for row_num, row_data in enumerate(table_data[1:], 2): #counting from second row -- skipping the first row for headers 
            for col_num, cell_data in enumerate(row_data, 1): #for each data row -- iterate through each column value -- assign a column number starting from 1
                cell = ws.cell(row=row_num, column=col_num)
                cell.value = cell_data #setting cell's content to data value

    wb.save(output_filename)
    print(f"Exceel file saved as: {output_filename}")

def main(pdf_path):
    try:
        images = convert_pdf_to_images(pdf_path)
        print(f"Successfully converted PDF to {len(images)} images")

    except Exception as e: #Exception is the general error type
        print(f"Error processing PDF: {e}")
        return None #when error occures we exit the function early, returning None

    '''
    "client" - basically opening a connection to google's servers
    client will send our images and recieve back ocr results
    '''
    client = vision.ImageAnnotatorClient() 
    all_text_data = [] #this empty list will store text data from each page as we process them

    for i, image in enumerate(images): #enumerate gives both image and page no(i). 
        print(f"Processing page {i+1}...")
        page_data = extract_text_from_image(client, image)
        all_text_data.append(page_data)

    table_data = [] #this will store our final table data - organized into rows and columns

    for page_data in all_text_data:
        structured_text = analyze_table_structure(page_data) #we process each page's ocr results through our analysis function
        table_data.extend(structured_text) #adding all our structured text blocks into one collection

    oraganized_table = organize_into_table(table_data) #positioned text block into organization function to a structured table
    print(f"Organized table with {len(oraganized_table)} rows") #user feedback: how many rows were detected
    
    for i, row in enumerate(oraganized_table[:5]):
        print(f"Row {i+1}: row") #displaying first 5 rows to verify table organisation

    output_filename = "extracted_table.xlsx" #definning the name of our output excel file
    write_table_to_excel(oraganized_table, output_filename) #calling our excel writing function with table data and desired filename
    print("PDF to Excel conversion completed successfully!")


if __name__=="__main__": #only run this code if someone executes this directly and not when imported as a module
    if len(sys.argv) != 2: #sys.argv contains command line arguments; index 0:script name, index 1:pdf file path
        print("Usage: python pdf_to_excel.py <path_to_pdf_file>") #if user didn't provide the right number of arguments 
        print("Example: python pdf_to_excel.py my_document.pdf")

        sys.exit()

    pdf_file_path = sys.argv[1] #we extract the PDF file path from command line arguments
    print(f"Starting PDF to Excel conversion for: {pdf_file_path}")

    main(pdf_file_path) #calling main function